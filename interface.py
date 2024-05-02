import os
import shutil
import sqlite3
import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import Frame
from ttkbootstrap import Style
from PIL import Image, ImageTk
from openpyxl.styles import Font
from tkinter import filedialog, StringVar

# Connect to SQLite database
conn = sqlite3.connect('entry_data.db')
cursor = conn.cursor()

# Create a table if it doesn't exist
cursor.execute('''CREATE TABLE IF NOT EXISTS EntryData (
                    id INTEGER PRIMARY KEY,
                    entry_text TEXT
                 )''')

# Define a function to save entry_var content into SQLite
def save_to_database(entry_text):
    try:
        # Create the table if it doesn't exist
        cursor.execute('''CREATE TABLE IF NOT EXISTS EntryData (
                            id INTEGER PRIMARY KEY,
                            entry_text TEXT
                         )''')
        
        # Drop existing data from the table
        cursor.execute("DELETE FROM EntryData")
        
        # Insert new data
        cursor.execute("INSERT INTO EntryData (entry_text) VALUES (?)", (entry_text,))
        conn.commit()
        print("Data saved to database successfully.")
        
        # Print the contents of the database
        print("Contents of the database:")
        cursor.execute("SELECT * FROM EntryData")
        for row in cursor.fetchall():
            print(row)
            
    except sqlite3.Error as e:
        print("Error saving data to database:", e)

# Define a function to handle confirmation of the input
def on_confirm(directory_path, input_dialog):
    input_dialog.destroy()  # Close the input dialog
    if directory_path:
        summary_file = create_summary_for_directory(directory_path)
        if summary_file:
            _, filename = os.path.split(summary_file)  # Extract just the filename
            label_text.config(text=f"Created successfully: {filename}")
            # Save entry_var content to database
            save_to_database(directory_path)
        else:
            label_text.config(text="Failed to create summary file.")

def create_summary(input_files, output_file):
    summary_workbook = openpyxl.Workbook()
    summary_sheet = summary_workbook.active
    blue_font = Font(color="0000FF")
    dark_green_font = Font(color="006400")

    summary_column = 1  # Start writing to the first column
    for input_file in input_files:
        try:
            input_workbook = openpyxl.load_workbook(input_file)
            input_sheet = input_workbook.active

            summary_row = 1  # Reset the row for each input file
            for row_idx, row in enumerate(input_sheet.iter_rows(values_only=True), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    # Write cells to the summary sheet in the appropriate column
                    summary_cell = summary_sheet.cell(row=summary_row, column=summary_column, value=cell)
                    if isinstance(cell, str):
                        if "PACKED" in cell.upper():
                            summary_cell.font = blue_font
                        elif cell and cell[0].isdigit():
                            summary_cell.font = dark_green_font
                    summary_row += 1
            summary_column += 1  # Move to the next column for the next input file
        except Exception as e:
            print(f"Error processing {input_file}: {e}")

    summary_workbook.save(output_file)
    print(f"Summary for {len(input_files)} Excel files created successfully!")
    return output_file  # Return the path of the created summary file

def create_summary_for_directory(directory):
    input_files = [os.path.join(directory, filename) for filename in os.listdir(directory) if filename.endswith(".xlsx")]
    output_file = os.path.join(directory, "jOkeo.xlsx")
    return create_summary(input_files, output_file)  # Return the path of the created summary file

def button1_clicked():
    label_text.config(text="Button 1 clicked")
    source_directory = filedialog.askdirectory(title="Select Source Directory")
    if source_directory:
        destination_directory = filedialog.askdirectory(title="Select Destination Directory")
        if destination_directory:
            copy_excel(source_directory, destination_directory)

def copy_excel(source_dir, destination_dir):
    if not os.path.exists(source_dir):
        label_text.config(text="Source directory not found")
        return
    
    if not os.path.exists(destination_dir):
        os.makedirs(destination_dir)

    total_copied = 0  # Initialize a counter for total copied files
    
    for root, dirs, files in os.walk(source_dir):
        for dir_name in dirs:
            subfolder_path = os.path.join(root, dir_name)
            excel_files = [f for f in os.listdir(subfolder_path) if f.endswith('.xlsx') or f.endswith('.xls')]
            if excel_files:
                for excel_file in excel_files:
                    source_file_path = os.path.join(subfolder_path, excel_file)
                    destination_file_path = os.path.join(destination_dir, excel_file)
                    if not os.path.exists(destination_file_path):
                        shutil.copy(source_file_path, destination_file_path)
                        total_copied += 1  # Increment the counter
                    else:
                        print(f"File '{excel_file}' already exists in '{destination_dir}'")

    label_text.config(text=f"Total quantity copied: {total_copied}")  # Update the label text

def button2_clicked():
    label_text.config(text="Kindly enter the address in the pop-up window.")
    
    # Create a Toplevel window for the input dialog
    input_dialog = tk.Toplevel(root)
    input_dialog.title("Enter Summary Location")

    # Set the size of the input dialog with a margin from the top
    input_dialog.geometry("300x150+100+100")  # Adjust the size and position as needed
    input_dialog.configure(pady=20)  # Add a margin from the top

    entry_var = StringVar()
    entry = tk.Entry(input_dialog, textvariable=entry_var, justify='center', font=('Helvetica', 12))  # You can adjust the font size here
    entry.pack(pady=10)  # Add some padding below the entry widget

    # Bind a function to clear the entry widget when it receives focus
    def clear_placeholder(event):
        if entry_var.get() == "Enter directory path...":
            entry_var.set("")

    # Set a placeholder text for the entry widget
    entry_var.set("Enter directory path...")

    # Bind the function to clear the placeholder text when the entry widget receives focus
    entry.bind("<FocusIn>", clear_placeholder)

    confirm_button = ttk.Button(input_dialog, text="Confirm", command=lambda: on_confirm(entry_var.get(), input_dialog), style="TButton.Outline")
    confirm_button.pack(pady=3)  # Add some padding above the button

def open_new_window():
    new_window = tk.Toplevel(root)
    new_window.title("View Shipped Date")
    new_window.geometry("340x360")  # Adjusted window size
    
    # Load the logo image
    logo_image_new = Image.open("C:/Users/joanm/Desktop/Noly/python 2/logo/logo2.png")
    logo_photo_new = ImageTk.PhotoImage(logo_image_new)

    # Save a reference to the image to prevent it from being garbage collected
    new_window.logo_photo = logo_photo_new

    # Create a frame for the logo
    logo_frame_new = Frame(new_window)
    logo_frame_new.pack()

    # Create a label to display the logo
    logo_label_new = tk.Label(logo_frame_new, image=logo_photo_new)
    logo_label_new.pack()

    # Add space between logo and input_entry
    tk.Label(new_window).pack()

    # Add a Text widget for displaying text
    text_widget = tk.Text(new_window, wrap='word', height=11, width=52)
    text_widget.pack(pady=(5, 10))

    # Define the action for the outline button directly
    def search_excel_file():
        try:
            # Clear the text widget
            text_widget.delete(1.0, tk.END)

            cursor.execute("SELECT entry_text FROM EntryData")
            directory_path = cursor.fetchone()[0]  # Assuming only one entry is stored in the database
            
            excel_file_path = os.path.join(directory_path, "jOkeo.xlsx")
            if os.path.exists(excel_file_path):
                # Load the Excel file
                workbook = openpyxl.load_workbook(excel_file_path)
                found = False
                result = []

                # Search for the entry_text in each sheet
                for sheet in workbook.sheetnames:
                    ws = workbook[sheet]
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.value and input_entry.get() in str(cell.value):
                                # If the search term is found, gather all values in the column
                                col_idx = cell.column
                                col_values = [row[col_idx - 1].value for row in ws.iter_rows(min_row=1) if row[col_idx - 1].value is not None]
                                result.append((col_idx, col_values))
                                found = True
                                break
                        if found:
                            break
                    if found:
                        break
                
                if not found:
                    text_widget.insert(tk.END, f"'{input_entry.get()}' not found in '{excel_file_path}'\n")
                else:
                    for idx, values in result:
                        text_widget.insert(tk.END, f"Column {idx}: {values}\n")
            else:
                text_widget.insert(tk.END, "File 'jOkeo.xlsx' not found in the directory.\n")
        except sqlite3.Error as e:
            print("Error retrieving data from database:", e)

    # Create the input_entry with placeholder text
    input_entry = tk.Entry(new_window, font=('Helvetica', 12), justify='center')
    input_entry.insert(0, "Type Serial Number")  # Add placeholder text
    input_entry.pack(pady=5)  # Add margin after the text widget

    # Define a function to clear the placeholder text when entry receives focus
    def clear_placeholder(event):
        if input_entry.get() == "Type Serial Number":
            input_entry.delete(0, tk.END)

    # Bind the function to clear the placeholder text
    input_entry.bind("<FocusIn>", clear_placeholder)

    # Create the search button
    search_button = ttk.Button(new_window, text="Search", style="TButton.Outline", command=search_excel_file)
    search_button.pack(pady=(5, 10))  # Add margin above and below the search button

    new_window.mainloop()

def button3_clicked():
    label_text.config(text="Did you find what your looking for?")
    open_new_window()

# Create a tkinter root window
root = tk.Tk()
root.title("Warranty Checker")

# Load logo
logo_image = Image.open("C:/Users/joanm/Desktop/Noly/python 2/logo/logo.png")
logo_photo = ImageTk.PhotoImage(logo_image)

# Create a frame for the logo
logo_frame = Frame(root)
logo_frame.pack()

# Create a label to display the logo
logo_label = tk.Label(logo_frame, image=logo_photo)
logo_label.pack()

# Create a frame for the buttons below the logo
button_frame = Frame(root)
button_frame.pack()

# Create a ttkbootstrap style
style = Style(theme='litera')
style.configure('TButton', font=('Helvetica', 10, 'bold'))  # <-- Add 'bold' to the font tuple

# Create buttons
button1 = ttk.Button(button_frame, text="Copy Excel", width=13, command=button1_clicked, style="TButton.Outline")
button2 = ttk.Button(button_frame, text="Create Summary", width=18, command=button2_clicked, style="TButton.Outline")
button3 = ttk.Button(button_frame, text="Search Serial", width=13, command=button3_clicked, style="TButton.Outline")

# Pack buttons in the button frame
button1.pack(side=tk.LEFT, padx=3, pady=10)
button2.pack(side=tk.LEFT, padx=3, pady=10)
button3.pack(side=tk.LEFT, padx=3, pady=10)

# Create a label for displaying text with a gray background
label_text = tk.Label(root, bg="white", width=50, height=5)
label_text.pack()

# Start the tkinter event loop
root.mainloop()

# Close database connection when the program exits
conn.close()
